#!/usr/bin/env python3
"""Convert the curated funding workbook into a readable, complete catalogue."""
from __future__ import annotations

import argparse
from datetime import date, datetime, timezone
import json
from pathlib import Path
import re

import openpyxl

ROOT = Path(__file__).resolve().parent


def funder(call: str) -> str:
    rules = [
        (r"\bFCT\b|Defense \+Science|Advanced Computing", "Fundação para a Ciência e a Tecnologia (FCT)"),
        (r"CMU Portugal", "CMU Portugal / FCT"),
        (r"UT Austin Portugal", "UT Austin Portugal / FCT"),
        (r"\bMIT\b|Massachusetts Institute of Technology", "MIT Portugal / FCT"),
        (r"Horizon|HORIZON", "European Commission — Horizon Europe"),
        (r"MSCA", "European Commission — Marie Skłodowska-Curie Actions"),
        (r"\bERC\b", "European Research Council"),
        (r"\bEIC\b", "European Innovation Council"),
        (r"Eurostars", "EUREKA / Eurostars"),
        (r"LIFE", "European Commission — LIFE Programme"),
        (r"ERDERA", "European Rare Diseases Research Alliance"),
        (r"Agroecology", "European Partnership Agroecology"),
        (r"Brain Health|EPBH", "European Partnership for Brain Health"),
        (r"CaixaImpulse", "la Caixa Foundation"),
        (r"Urban Mobility", "EIT Urban Mobility"),
        (r"Olympic|Paralympic", "Portuguese Olympic and Paralympic Committees"),
        (r"Bluepharma", "Bluepharma / University of Coimbra"),
        (r"PRIMA", "PRIMA Foundation / FCT"),
        (r"Air Traffic", "Air navigation competition — organizer not stated"),
        (r"INOVA\+", "INOVA+"),
        (r"J\. Norberto Pires", "University of Coimbra — J. Norberto Pires Prize"),
        (r"GENESIS", "University of Coimbra — FCTUC"),
    ]
    return next((name for pattern, name in rules if re.search(pattern, call, re.I)), "Organizer not stated in workbook")


def priority(call: str, info: str) -> tuple[str, str]:
    text = f"{call} {info}".lower()
    high = ("artificial intelligence", "robot", "digital", "advanced computing", "computer", "information and communication", "deep reasoning", "manufacturing", "eurostars", "eic pathfinder")
    medium = ("all areas", "mobility", "energy", "climate", "agro", "agricultur", "defense", "security", "space", "materials", "researchers", "ph.d", "scientific employment")
    if any(term in text for term in high) or re.search(r"\bai\b", text):
        return "High", "Direct AI, robotics, computing, digital, or technology-development fit"
    if any(term in text for term in medium):
        return "Medium", "Broad or adjacent fit requiring a suitable robotics/AI application"
    return "Low", "Sector-specific or weak fit unless a targeted collaboration is available"


def clean(value) -> str:
    return str(value or "").replace("\n", "<br>").replace("|", "\\|").strip()


def load(workbook: Path) -> list[dict]:
    book = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    sheet = book["2026"] if "2026" in book.sheetnames else book.active
    rows = []
    for values in sheet.iter_rows(min_row=3, values_only=True):
        deadline, call, info = (tuple(values) + (None, None, None))[:3]
        if not call:
            continue
        if isinstance(deadline, datetime):
            deadline = deadline.date()
        elif not isinstance(deadline, date):
            deadline = datetime.fromisoformat(str(deadline)).date()
        level, reason = priority(str(call), str(info or ""))
        rows.append({"deadline": deadline, "call": str(call), "info": str(info or ""), "funder": funder(str(call)), "priority": level, "reason": reason})
    return sorted(rows, key=lambda row: (row["deadline"], row["call"]))


def load_verified(path: Path) -> list[dict]:
    data = json.loads(path.read_text(encoding="utf-8"))
    return sorted(data.get("calls", []), key=lambda item: (item["deadline"], item["program"]))


def load_workbook_links(path: Path) -> dict[str, dict]:
    return json.loads(path.read_text(encoding="utf-8"))


def render(rows: list[dict], as_of: date, verified: list[dict], links: dict[str, dict]) -> str:
    future = [row for row in rows if row["deadline"] >= as_of]
    past = [row for row in rows if row["deadline"] < as_of]
    high = sum(row["priority"] == "High" for row in future)
    medium = sum(row["priority"] == "Medium" for row in future)
    lines = [
        "# Funding Sources Catalogue",
        "",
        f"**Workbook:** `funding_calls.xlsx`",
        f"**Analysed:** {as_of.isoformat()}",
        f"**Coverage:** {len(rows)} calls ({len(future)} open/future by deadline; {len(past)} past)",
        f"**Official links:** {sum(row['call'] in links for row in future)} of {len(future)} open/future workbook entries resolved to official sources; unresolved entries remain explicitly marked pending.",
        "",
        "## Analysis",
        "",
        f"Among the {len(future)} non-past calls, {high} are high-priority and {medium} are medium-priority for ISR/UC work in AI, robotics, electrical/computer engineering, perception, autonomy, or related applications. High priority means direct thematic alignment; medium priority requires positioning the technical work inside a broader domain.",
        "",
    ]
    lines += ["## Funding bodies represented", "", "| Funding source | Total calls | Open/future |", "|---|---:|---:|"]
    bodies = sorted({row["funder"] for row in rows} | {item["funder"] for item in verified})
    for body in bodies:
        total = sum(row["funder"] == body for row in rows) + sum(item["funder"] == body for item in verified)
        active = sum(row["funder"] == body for row in future) + sum(item["funder"] == body and date.fromisoformat(item["deadline"]) >= as_of for item in verified)
        lines.append(f"| {clean(body)} | {total} | {active} |")
    lines += ["", "## Newly verified from official portals", "", "| Deadline | Call | Funding source | Priority | Scope | Official source |", "|---|---|---|---|---|---|"]
    for item in verified:
        lines.append(f"| {item['deadline']} | {clean(item['program'])} | {clean(item['funder'])} | **{item['priority']}** | {clean(item['objective'])} | [Official page]({item['link']}) |")
    lines += ["", "## Open and future workbook calls", "", "| Deadline | Call | Funding source (inferred) | Priority | Rationale | Additional information | Official source |", "|---|---|---|---|---|---|---|"]
    for row in future:
        source = links.get(row["call"])
        official = f"[Official page]({source['url']})<br>{clean(source['note'])}" if source else "**Official notice not located**"
        lines.append(f"| {row['deadline']} | {clean(row['call'])} | {clean(row['funder'])} | **{row['priority']}** | {clean(row['reason'])} | {clean(row['info']) or '—'} | {official} |")
    lines += ["", "## Past calls retained for source history", "", "| Deadline | Call | Funding source (inferred) | Priority | Additional information |", "|---|---|---|---|---|"]
    for row in past:
        lines.append(f"| {row['deadline']} | {clean(row['call'])} | {clean(row['funder'])} | {row['priority']} | {clean(row['info']) or '—'} |")
    lines += ["", "## Maintenance notes", "", "- Preserve past calls as a source history; do not present them as active.", "- Add official URLs only after verifying them against the funding body's portal.", "- Re-run `python3 funding-radar/import_funding_sources.py` after editing the workbook.", "- Priority is a screening aid, not an eligibility decision.", ""]
    return "\n".join(lines)


def dashboard_call(row: dict, as_of: date, links: dict[str, dict]) -> dict:
    days = (row["deadline"] - as_of).days
    status = "🔴" if days <= 30 else "🟡" if days <= 90 else "🟢"
    info = row["info"] or "See official call conditions"
    support_match = re.search(r"(?:co-?funding(?: of)?|funding)\s*(?:of\s*)?(\d+%)", info, re.I)
    budget_match = re.search(r"(?:€|€\s*)([\d.,]+(?:\s*[kKmM])?)", info)
    source = links.get(row["call"], {})
    note = source.get("note", "")
    objective = f"{info} Verification note: {note}" if "unverified" in note else info
    return {
        "program": row["call"],
        "deadline": row["deadline"].strftime("%b %d, %Y"),
        "budget": f"€{budget_match.group(1)}" if budget_match else "See call",
        "support": support_match.group(1) if support_match else "See call",
        "eligible": "Verify in official call",
        "objective": objective,
        "fit": f"{row['priority']} (workbook screening)",
        "link": source.get("url", ""),
        "status": status,
        "source": "funding_calls.xlsx",
    }


def update_dashboard(rows: list[dict], as_of: date, calls_path: Path, verified: list[dict], links: dict[str, dict]) -> None:
    data = json.loads(calls_path.read_text(encoding="utf-8"))
    superseded = {"ERC Starting Grants", "KDT Joint Undertaking"}
    retained = [call for call in data.get("calls", []) if call.get("source") not in ("funding_calls.xlsx", "verified_official_scan") and call.get("program") not in superseded]
    workbook_calls = [dashboard_call(row, as_of, links) for row in rows if row["deadline"] >= as_of]
    verified_calls = [{"program": item["program"], "deadline": datetime.fromisoformat(item["deadline"]).strftime("%b %d, %Y"), "budget": item["budget"], "support": item["support"], "eligible": item["eligible"], "objective": item.get("why_now", item["objective"]), "fit": f"{item['fit_score']}/5 {item['priority']}" if "fit_score" in item else item["priority"], "link": item["link"], "status": "🔴" if (date.fromisoformat(item["deadline"]) - as_of).days <= 30 else "🟡" if (date.fromisoformat(item["deadline"]) - as_of).days <= 90 else "🟢", "source": "verified_official_scan"} for item in verified if date.fromisoformat(item["deadline"]) >= as_of]
    data["calls"] = retained + workbook_calls + verified_calls
    data["timestamp"] = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    calls_path.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=ROOT / "funding_calls.xlsx")
    parser.add_argument("--output", type=Path, default=ROOT / "FUNDING_SOURCES.md")
    parser.add_argument("--calls", type=Path, default=ROOT / "calls.json")
    parser.add_argument("--verified", type=Path, default=ROOT / "verified_calls.json")
    parser.add_argument("--workbook-links", type=Path, default=ROOT / "workbook_call_links.json")
    parser.add_argument("--as-of", type=date.fromisoformat, default=date.today())
    args = parser.parse_args()
    rows = load(args.workbook)
    verified = load_verified(args.verified)
    links = load_workbook_links(args.workbook_links)
    args.output.write_text(render(rows, args.as_of, verified, links), encoding="utf-8")
    update_dashboard(rows, args.as_of, args.calls, verified, links)
    print(f"wrote {args.output} with {len(rows)} calls")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
